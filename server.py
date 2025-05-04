from flask import Flask, request, jsonify, render_template, send_from_directory
import pandas as pd
import os
from datetime import datetime
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import uuid
from werkzeug.utils import secure_filename

app = Flask(__name__, static_folder='.', static_url_path='')

# Create student_data directory if it doesn't exist
os.makedirs('student_data', exist_ok=True)

# Create directory for photo submissions
PHOTO_UPLOADS_DIR = 'images/user_submissions'
os.makedirs(PHOTO_UPLOADS_DIR, exist_ok=True)

# Define the master Excel file path
MASTER_EXCEL_FILE = 'student_data/all_admissions.xlsx'

@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    return send_from_directory('.', path)

def initialize_excel_file():
    """Create or initialize the master Excel file if it doesn't exist"""
    if not os.path.exists(MASTER_EXCEL_FILE):
        # Create a new workbook
        wb = openpyxl.Workbook()
        
        # Create separate sheets for different program types
        wb.active.title = "Intermediate Admissions"
        wb.create_sheet("BS Admissions")
        wb.create_sheet("MS Admissions")
        wb.create_sheet("Summary")
        
        # Set up headers for each sheet
        sheets = ["Intermediate Admissions", "BS Admissions", "MS Admissions"]
        
        for sheet_name in sheets:
            sheet = wb[sheet_name]
            headers = [
                "Submission ID", "Submission Date", "Full Name", "Father Name", 
                "Gender", "Date of Birth", "CNIC", "Father CNIC", "Program",
                "Previous Education", "Previous Marks", "Previous Total", "Percentage",
                "Address", "City", "Email", "Phone", "Guardian Phone", "Documents Submitted"
            ]
            
            # Add headers to the first row
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4e008e", end_color="4e008e", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # Adjust column widths
            for col_num in range(1, len(headers) + 1):
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
        
        # Set up summary sheet
        summary_sheet = wb["Summary"]
        summary_headers = ["Program", "Total Applications", "Male Applicants", "Female Applicants", "Last Updated"]
        
        for col_num, header in enumerate(summary_headers, 1):
            cell = summary_sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4e008e", end_color="4e008e", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # Set column widths for summary
        for col_num in range(1, len(summary_headers) + 1):
            summary_sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
            
        # Add program categories to summary
        programs = ["F.Sc (Pre-Medical)", "F.Sc (Pre-Engineering)", "ICS", "I.Com", "F.A", 
                   "BS Computer Science", "BS Mathematics", "BS Physics", "BS Chemistry", "BS English",
                   "MS Computer Science", "MS Mathematics", "MS Physics", "Others"]
                   
        for row_num, program in enumerate(programs, 2):
            summary_sheet.cell(row=row_num, column=1).value = program
            summary_sheet.cell(row=row_num, column=2).value = 0
            summary_sheet.cell(row=row_num, column=3).value = 0
            summary_sheet.cell(row=row_num, column=4).value = 0
            summary_sheet.cell(row=row_num, column=5).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
        # Add total row at the bottom
        total_row = len(programs) + 2
        summary_sheet.cell(row=total_row, column=1).value = "TOTAL"
        summary_sheet.cell(row=total_row, column=1).font = Font(bold=True)
        
        # Add formulas for totals
        summary_sheet.cell(row=total_row, column=2).value = f"=SUM(B2:B{total_row-1})"
        summary_sheet.cell(row=total_row, column=3).value = f"=SUM(C2:C{total_row-1})"
        summary_sheet.cell(row=total_row, column=4).value = f"=SUM(D2:D{total_row-1})"
        summary_sheet.cell(row=total_row, column=5).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Save the workbook
        wb.save(MASTER_EXCEL_FILE)
        print(f"Master Excel file created at {MASTER_EXCEL_FILE}")

def update_excel_with_submission(form_data):
    """Update the master Excel file with a new submission"""
    # Make sure the Excel file exists
    initialize_excel_file()
    
    # Load the workbook
    wb = openpyxl.load_workbook(MASTER_EXCEL_FILE)
    
    # Determine which sheet to use based on program level
    program_level = form_data.get('programLevel', '').lower()
    if program_level == 'intermediate':
        sheet_name = "Intermediate Admissions"
    elif program_level == 'bachelor':
        sheet_name = "BS Admissions"
    elif program_level == 'masters':
        sheet_name = "MS Admissions"
    else:
        sheet_name = "Intermediate Admissions"  # Default
    
    sheet = wb[sheet_name]
    
    # Find the next available row
    next_row = sheet.max_row + 1
    
    # Add a unique submission ID 
    submission_id = f"{sheet_name[0]}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    
    # Map form data to columns
    data = [
        submission_id,
        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        form_data.get('fullname', ''),
        form_data.get('fathername', ''),
        form_data.get('gender', ''),
        form_data.get('dob', ''),
        form_data.get('cnic', ''),
        form_data.get('fatherCnic', ''),
        form_data.get('program', ''),
        form_data.get('previousEducation', ''),
        form_data.get('previousMarks', ''),
        form_data.get('previousTotal', ''),
        form_data.get('percentage', ''),
        form_data.get('address', ''),
        form_data.get('city', ''),
        form_data.get('email', ''),
        form_data.get('phone', ''),
        form_data.get('guardianPhone', ''),
        ', '.join(form_data.get('documents', []))
    ]
    
    # Write data to the row
    for col_num, value in enumerate(data, 1):
        sheet.cell(row=next_row, column=col_num).value = value
    
    # Update summary sheet
    update_summary(wb, form_data)
    
    # Save the workbook
    wb.save(MASTER_EXCEL_FILE)
    
    return submission_id

def update_summary(wb, form_data):
    """Update the summary sheet with the new submission"""
    summary_sheet = wb["Summary"]
    
    # Map program to summary row
    program_map = {
        "fsc-premedical": "F.Sc (Pre-Medical)",
        "fsc-preengineering": "F.Sc (Pre-Engineering)",
        "ics": "ICS",
        "icom": "I.Com",
        "fa": "F.A",
        "bs-cs": "BS Computer Science",
        "bs-math": "BS Mathematics",
        "bs-physics": "BS Physics",
        "bs-chemistry": "BS Chemistry",
        "bs-english": "BS English",
        "ms-cs": "MS Computer Science",
        "ms-math": "MS Mathematics",
        "ms-physics": "MS Physics"
    }
    
    program = form_data.get('program', '')
    summary_program = program_map.get(program, "Others")
    
    # Find the row for this program
    program_row = None
    for row in range(2, summary_sheet.max_row):
        if summary_sheet.cell(row=row, column=1).value == summary_program:
            program_row = row
            break
    
    if program_row:
        # Increment total applications
        current_count = summary_sheet.cell(row=program_row, column=2).value or 0
        summary_sheet.cell(row=program_row, column=2).value = current_count + 1
        
        # Update gender counts
        gender = form_data.get('gender', '').lower()
        if gender == 'male':
            current_male = summary_sheet.cell(row=program_row, column=3).value or 0
            summary_sheet.cell(row=program_row, column=3).value = current_male + 1
        elif gender == 'female':
            current_female = summary_sheet.cell(row=program_row, column=4).value or 0
            summary_sheet.cell(row=program_row, column=4).value = current_female + 1
        
        # Update timestamp
        summary_sheet.cell(row=program_row, column=5).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

@app.route('/api/submit-admission', methods=['POST'])
def submit_admission():
    try:
        # Get form data as JSON
        form_data = request.json
        
        # Add timestamp
        form_data['submission_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Log the received data for debugging
        print("Received form data:", form_data)
        
        # Update the master Excel file
        submission_id = update_excel_with_submission(form_data)
        
        # Also save individual file as JSON for backup
        student_name = form_data.get('fullname', 'unknown').replace(' ', '_')
        program = form_data.get('program', 'unknown')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        json_filename = f"student_data/{student_name}_{program}_{timestamp}.json"
        
        with open(json_filename, 'w') as f:
            json.dump(form_data, f, indent=4)
        
        # Return success response
        return jsonify({
            'success': True,
            'message': 'Application submitted successfully',
            'submission_id': submission_id
        })
    
    except Exception as e:
        # Log the error
        print(f"Error processing admission form: {str(e)}")
        
        # Return error response
        return jsonify({
            'success': False,
            'message': f'Error processing your application: {str(e)}'
        }), 500

@app.route('/api/list-admissions', methods=['GET'])
def list_admissions():
    try:
        # Get list of Excel files in student_data directory
        files = [f for f in os.listdir('student_data') if f.endswith('.xlsx')]
        
        # Extract student information from filenames
        students = []
        for file in files:
            parts = file.split('_')
            if len(parts) >= 3:
                name = parts[0].replace('_', ' ')
                program = parts[1]
                date = '_'.join(parts[2:]).replace('.xlsx', '')
                
                students.append({
                    'name': name,
                    'program': program,
                    'date': date,
                    'filename': file
                })
        
        # Return the list
        return jsonify({
            'success': True,
            'admissions': students
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error retrieving admissions: {str(e)}'
        }), 500

@app.route('/api/download-admission/<filename>', methods=['GET'])
def download_admission(filename):
    return send_from_directory('student_data', filename)

@app.route('/api/submit-photo', methods=['POST'])
def submit_photo():
    try:
        # Check if the post request has the file part
        if 'photo-file' not in request.files:
            return jsonify({
                'success': False,
                'message': 'No file part in the request'
            }), 400
            
        file = request.files['photo-file']
        
        # If user does not select file, browser also
        # submit an empty part without filename
        if file.filename == '':
            return jsonify({
                'success': False,
                'message': 'No selected file'
            }), 400
            
        # Generate unique filename to avoid collisions
        filename = secure_filename(file.filename)
        unique_id = str(uuid.uuid4())
        filename_parts = os.path.splitext(filename)
        unique_filename = f"{filename_parts[0]}_{unique_id}{filename_parts[1]}"
        
        # Save the file
        file_path = os.path.join(PHOTO_UPLOADS_DIR, unique_filename)
        file.save(file_path)
        
        # Get form data
        submitter_name = request.form.get('submitter-name', 'Anonymous')
        submitter_email = request.form.get('submitter-email', '')
        submitter_role = request.form.get('submitter-role', '')
        photo_title = request.form.get('photo-title', 'Untitled')
        photo_description = request.form.get('photo-description', '')
        photo_category = request.form.get('photo-category', '')
        photo_date = request.form.get('photo-date', '')
        
        # Save metadata as JSON
        metadata = {
            'filename': unique_filename,
            'original_filename': filename,
            'submitter_name': submitter_name,
            'submitter_email': submitter_email,
            'submitter_role': submitter_role,
            'title': photo_title,
            'description': photo_description,
            'category': photo_category,
            'date_taken': photo_date,
            'submission_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'status': 'pending'  # This could be 'pending', 'approved', 'rejected'
        }
        
        # Save metadata to a JSON file
        metadata_filename = os.path.join(PHOTO_UPLOADS_DIR, f"{unique_id}.json")
        with open(metadata_filename, 'w') as f:
            json.dump(metadata, f, indent=4)
        
        # Log the submission
        print(f"Photo submitted: {unique_filename} by {submitter_name}")
        
        return jsonify({
            'success': True,
            'message': 'Photo submitted successfully! It will be reviewed before being published.',
            'filename': unique_filename,
        })
        
    except Exception as e:
        # Log the error
        print(f"Error processing photo submission: {str(e)}")
        
        # Return error response
        return jsonify({
            'success': False,
            'message': f'Error processing your photo: {str(e)}'
        }), 500

@app.route('/api/get-gallery-photos', methods=['GET'])
def get_gallery_photos():
    """Return a list of all approved user-submitted photos"""
    try:
        gallery_photos = []
        
        # Get all JSON files in the uploads directory
        json_files = [f for f in os.listdir(PHOTO_UPLOADS_DIR) if f.endswith('.json')]
        
        for json_file in json_files:
            json_path = os.path.join(PHOTO_UPLOADS_DIR, json_file)
            with open(json_path, 'r') as f:
                metadata = json.load(f)
                
            # Only include approved photos (in real implementation)
            # For now, include all photos for demonstration
            photo_path = os.path.join(PHOTO_UPLOADS_DIR, metadata['filename'])
            if os.path.exists(photo_path):
                gallery_photos.append({
                    'filename': metadata['filename'],
                    'title': metadata['title'],
                    'description': metadata['description'],
                    'category': metadata['category'],
                    'submitter_name': metadata['submitter_name'],
                    'submission_date': metadata['submission_date'],
                    'url': f"/images/user_submissions/{metadata['filename']}"
                })
        
        return jsonify({
            'success': True,
            'photos': gallery_photos
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error retrieving gallery photos: {str(e)}'
        }), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)